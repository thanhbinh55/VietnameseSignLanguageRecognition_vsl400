"""Bộ_Cập_Nhật_Nhãn — đọc bảng nhãn nguồn và xuất bảng nhãn mới (Requirement 7).

Module này định nghĩa:

* :class:`LabelRow` — một dòng nhãn với cấu trúc cột QIPEDC
  ``STT, ID, VIDEO, LABEL, REGION, TOPIC, signer``.
* :func:`read_source_labels` — đọc ``Dataset/labels/*.xlsx`` bằng :mod:`openpyxl`,
  khớp header **không phân biệt hoa/thường** (lớp biên I/O).
* :func:`build_new_rows` — LOGIC THUẦN dựng các dòng nhãn mới: mỗi
  :class:`~qipedc_video_preprocess.splitter.OutputClip` sinh **một** dòng kế thừa
  ``ID/LABEL/REGION/TOPIC/signer`` của dòng nguồn cùng ``video_id`` với ``VIDEO``
  bằng ``out_filename``; bỏ qua dòng nguồn trỏ video ngoài tập đã duyệt; gán lại
  ``STT`` thành dãy ``1..M`` duy nhất tăng dần.
* :func:`write_new_labels` — ghi ``.xlsx`` mới tại ``cfg.new_labels_path`` (khác
  và không ghi đè Bảng_Nhãn_Nguồn); không có dòng dữ liệu → chỉ ghi header.

Xem design.md, Requirement 7.
"""

from __future__ import annotations

import glob
import logging
import re
from dataclasses import dataclass, replace
from pathlib import Path
from typing import Any, Iterable, Sequence

import openpyxl

logger = logging.getLogger(__name__)

# Cấu trúc cột chuẩn của bảng nhãn QIPEDC, theo đúng thứ tự (Req 7.3). Tên cột
# được khớp không phân biệt hoa/thường khi đọc; khi ghi, header dùng đúng nhãn
# hiển thị dưới đây.
COLUMN_HEADERS: tuple[str, ...] = (
    "STT",
    "ID",
    "VIDEO",
    "LABEL",
    "REGION",
    "TOPIC",
    "signer",
)
_FIELDS: tuple[str, ...] = ("stt", "id", "video", "label", "region", "topic", "signer")

# Chuỗi thuần số nguyên/thập phân (vd "1", "1.0") — để chuẩn hóa nhãn số float
# của Excel ("1.0" -> "1") mà giữ nguyên từ thật / tiếng Việt có dấu.
_FLOAT_RE = re.compile(r"^[+-]?\d+(\.\d+)?$")


@dataclass(frozen=True)
class LabelRow:
    """Một dòng nhãn QIPEDC (Req 7).

    Mọi trường là ``str | None`` vì ô bảng tính có thể trống. Thứ tự cột phản ánh
    :data:`COLUMN_HEADERS`.
    """

    stt: str | None
    id: str | None
    video: str | None
    label: str | None
    region: str | None
    topic: str | None
    signer: str | None


# --------------------------------------------------------------------------- #
# Tiện ích chuẩn hóa ô (lớp biên đọc)
# --------------------------------------------------------------------------- #
def _num_to_str(value: int | float) -> str:
    """Render số thành chuỗi chuẩn: float nguyên mất ``.0`` (``1.0`` -> ``"1"``)."""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return repr(value)
    return str(value)


def _coerce_cell(value: Any) -> str | None:
    """Chuyển ô openpyxl bất kỳ thành chuỗi sạch hoặc ``None`` (ô trống)."""
    if value is None:
        return None
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (int, float)):
        return _num_to_str(value)
    text = str(value).strip()
    return text or None


def _normalize_label(value: Any) -> str | None:
    """Chuẩn hóa giá trị ``LABEL``: nhãn số float ``1.0`` -> ``"1"``; từ thật giữ nguyên."""
    if value is None:
        return None
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (int, float)):
        return _num_to_str(value)
    text = str(value).strip()
    if not text:
        return None
    if _FLOAT_RE.match(text):
        return _num_to_str(float(text))
    return text


def video_id_of(video: str | None) -> str | None:
    """Suy ra ``video_id`` (stem) từ giá trị cột ``VIDEO`` (vd ``"W00202.mp4"`` -> ``"W00202"``).

    Trả về ``None`` nếu ``video`` rỗng/None. Việc khớp dùng stem nên ``VIDEO`` có
    hay không có đuôi ``.mp4`` đều quy về cùng ``video_id``.
    """
    if not video:
        return None
    return Path(str(video).strip()).stem or None


# --------------------------------------------------------------------------- #
# Đọc bảng nhãn nguồn (BIÊN — openpyxl)
# --------------------------------------------------------------------------- #
def _label_files(cfg) -> list[str]:
    """Danh sách ``*.xlsx`` khớp ``cfg.labels_glob`` (đã sắp xếp, tái lập được)."""
    pattern = str(Path(cfg.project_root) / cfg.labels_glob)
    return sorted(glob.glob(pattern))


def _build_header_map(header_row: "tuple[Any, ...] | None") -> dict[str, int]:
    """Ánh xạ tên cột QIPEDC -> chỉ số cột, khớp không phân biệt hoa/thường."""
    mapping: dict[str, int] = {}
    if not header_row:
        return mapping
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        name = str(cell).strip().lower()
        if name in _FIELDS and name not in mapping:
            mapping[name] = idx
    return mapping


def _is_blank_row(row: "tuple[Any, ...] | None") -> bool:
    if not row:
        return True
    return all(
        cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row
    )


def _row_from_cells(row: "tuple[Any, ...]", header_map: dict[str, int]) -> LabelRow:
    def get(field: str) -> Any:
        idx = header_map.get(field)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    return LabelRow(
        stt=_coerce_cell(get("stt")),
        id=_coerce_cell(get("id")),
        video=_coerce_cell(get("video")),
        label=_normalize_label(get("label")),
        region=_coerce_cell(get("region")),
        topic=_coerce_cell(get("topic")),
        signer=_coerce_cell(get("signer")),
    )


def _read_one_file(path: str) -> list[LabelRow]:
    """Đọc một file ``.xlsx`` thành danh sách :class:`LabelRow`."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows: list[LabelRow] = []
        header_map: dict[str, int] | None = None
        for row in worksheet.iter_rows(values_only=True):
            if header_map is None:
                if _is_blank_row(row):
                    continue
                header_map = _build_header_map(row)
                continue
            if _is_blank_row(row):
                continue
            rows.append(_row_from_cells(row, header_map))
        return rows
    finally:
        workbook.close()


def read_source_labels(cfg) -> list[LabelRow]:
    """Đọc toàn bộ bảng nhãn nguồn khớp ``cfg.labels_glob`` (Req 7).

    Gộp tất cả file ``*.xlsx`` (sắp xếp để tái lập), khớp header không phân biệt
    hoa/thường với ``STT, ID, VIDEO, LABEL, REGION, TOPIC, signer``. Nhãn số float
    được chuẩn hóa qua :func:`_normalize_label`.

    Args:
        cfg: :class:`~qipedc_video_preprocess.config.PreprocessConfig`.

    Returns:
        ``list[LabelRow]`` phẳng trên tất cả file.
    """
    rows: list[LabelRow] = []
    for path in _label_files(cfg):
        rows.extend(_read_one_file(path))
    return rows


# --------------------------------------------------------------------------- #
# Dựng dòng nhãn mới (LOGIC THUẦN — Req 7.1/7.2/7.3/7.5)
# --------------------------------------------------------------------------- #
def build_new_rows(
    source_rows: "Sequence[LabelRow]",
    output_clips: "Iterable",
    discovered_ids: "Iterable[str]",
    logger=logger,
) -> list[LabelRow]:
    """Dựng các dòng nhãn mới từ các clip đầu ra (THUẦN — Req 7.1/7.2/7.3/7.5).

    Với mỗi :class:`~qipedc_video_preprocess.splitter.OutputClip`:

    * Sinh **đúng một** :class:`LabelRow` với ``VIDEO = clip.out_filename`` và kế
      thừa ``ID/LABEL/REGION/TOPIC/signer`` của dòng nguồn cùng ``video_id``
      (Req 7.1/7.2).
    * Nếu ``clip.video_id`` không thuộc ``discovered_ids`` thì **bỏ qua** clip đó
      và ghi log ``video_id`` thiếu (Req 7.5).
    * Nếu không tìm thấy dòng nguồn cho ``video_id`` (clip ứng với video đã duyệt
      nhưng không có dòng nhãn nguồn), các trường kế thừa để ``None``.

    Sau cùng gán lại cột ``STT`` thành dãy ``1..M`` duy nhất, tăng dần liên tục
    (Req 7.3).

    Args:
        source_rows: Các :class:`LabelRow` nguồn (từ :func:`read_source_labels`).
        output_clips: Iterable các ``OutputClip`` đã lên kế hoạch.
        discovered_ids: Tập các ``video_id`` đã duyệt được (để lọc Req 7.5).
        logger: Logger để ghi ``video_id`` bị thiếu.

    Returns:
        ``list[LabelRow]`` mới với ``STT`` đã đánh lại ``1..M``.
    """
    discovered = set(discovered_ids)

    # Map video_id -> dòng nguồn (giữ dòng đầu tiên cho mỗi video_id).
    source_by_id: dict[str, LabelRow] = {}
    for row in source_rows:
        vid = video_id_of(row.video)
        if vid is not None and vid not in source_by_id:
            source_by_id[vid] = row

    new_rows: list[LabelRow] = []
    for clip in output_clips:
        vid = clip.video_id
        if vid not in discovered:
            logger.warning(
                "build_new_rows: bỏ qua clip %s — video_id %r không thuộc tập đã duyệt",
                getattr(clip, "out_filename", "?"),
                vid,
            )
            continue

        src = source_by_id.get(vid)
        new_rows.append(
            LabelRow(
                stt=None,  # đánh lại bên dưới
                id=src.id if src is not None else None,
                video=clip.out_filename,
                label=src.label if src is not None else None,
                region=src.region if src is not None else None,
                topic=src.topic if src is not None else None,
                signer=src.signer if src is not None else None,
            )
        )

    # Gán lại STT = 1..M duy nhất tăng dần liên tục (Req 7.3).
    return [replace(row, stt=str(i)) for i, row in enumerate(new_rows, start=1)]


# --------------------------------------------------------------------------- #
# Ghi bảng nhãn mới (BIÊN — openpyxl, Req 7.4/7.7)
# --------------------------------------------------------------------------- #
def write_new_labels(new_rows: "Sequence[LabelRow]", cfg, logger=logger) -> Path:
    """Ghi Bảng_Nhãn_Mới ra ``.xlsx`` tại ``cfg.new_labels_path`` (Req 7.4/7.7).

    Đường dẫn đầu ra **khác** mọi Bảng_Nhãn_Nguồn nên không ghi đè nguồn (Req 7.4
    — ``validate()`` của config đã bảo đảm ``new_labels_path`` nằm trong cây dự án
    trên ``D:`` và mặc định trỏ vào ``split_variants/``, tách khỏi
    ``Dataset/labels/``). Khi ``new_rows`` rỗng, vẫn ghi file chỉ gồm **dòng
    header** với đúng cấu trúc cột và ghi log (Req 7.7).

    Args:
        new_rows: Các :class:`LabelRow` đầu ra (từ :func:`build_new_rows`).
        cfg: :class:`~qipedc_video_preprocess.config.PreprocessConfig`.
        logger: Logger.

    Returns:
        Đường dẫn :class:`Path` của file đã ghi.
    """
    out_path = cfg.resolve(cfg.new_labels_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(list(COLUMN_HEADERS))

    for row in new_rows:
        worksheet.append(
            [row.stt, row.id, row.video, row.label, row.region, row.topic, row.signer]
        )

    workbook.save(str(out_path))
    workbook.close()

    if not new_rows:
        logger.warning(
            "write_new_labels: không có dòng nhãn nào — ghi %s chỉ gồm header",
            out_path.name,
        )

    return out_path
