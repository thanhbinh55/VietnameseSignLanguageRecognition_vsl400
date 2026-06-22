"""Cắt_Tay — cắt video theo điểm tự nhập (CSV/XLSX) cho nhiều cách / nhiều góc.

Công cụ này KHÔNG dùng OCR: người dùng tự xác định điểm cắt và khai báo trong một
bảng CSV/XLSX, công cụ cắt video thành các đoạn và đặt tên theo đúng quy ước của
pipeline (``_c1/_c2/_c3`` cho nhiều cách; ``_side`` cho góc phụ). Dùng cho các
video mà bộ phát hiện tự động không xử lý được (gom trong folder ``manual``), hoặc
khi muốn cắt chính xác theo mốc thời gian đã biết.

Tái sử dụng các khối đã được kiểm thử của pipeline:
:func:`splitter.write_clip`, :func:`splitter.variant_filename`,
:func:`splitter.view_filename`, :class:`segmenter.VariantSpan`,
:func:`video_probe.probe`, :func:`discovery.discover_videos`.

Lược đồ bảng (CSV hoặc XLSX), header không phân biệt hoa/thường:

============== ==============================================================
Cột            Ý nghĩa
============== ==============================================================
``video_id``   stem video nguồn, vd ``W00738``
``mode``       ``multiway`` | ``multiview`` | ``both``
``cut_seconds``     mốc bắt đầu (giây) của cách 2,3… — vd ``2.5`` hoặc ``2.5,5.0``
                    (dùng cho ``multiway`` và ``both``)
``view_cut_seconds`` mốc cắt cứng front→side (giây). ``multiview``: một mốc cho
                    cả video. ``both``: một mốc cho MỖI cách (số mốc = số cách),
                    là thời điểm tuyệt đối trong video gốc.
``notes``      ghi chú tự do (không bắt buộc)
============== ==============================================================

Quy ước đặt tên:

* ``multiway`` (N cách): ``<id>_c1.mp4 … _cN.mp4``.
* ``multiview``: đoạn trước (front) giữ ``<id>.mp4``; đoạn sau ``<id>_side.mp4``.
* ``both`` (METHOD-MAJOR): cắt theo cách trước (``_c1/_c2``); trong MỖI cách lại
  cắt theo góc → ``<id>_cK.mp4`` (front) + ``<id>_cK_side.mp4`` (side).

Clip được ghi vào ``cfg.split_output_path`` (split_variants/) — đây là kết quả đã
được người duyệt xác nhận. Cắt theo mốc CHÍNH XÁC do người dùng nhập (KHÔNG trừ
Biên_An_Toàn, không suy đoán).

CLI::

    python -m qipedc_video_preprocess.manual_cut \
        --input Dataset/processed_videos/manual_cuts.csv [--project-root .] [--dry-run]
"""

from __future__ import annotations

import argparse
import csv
import logging
import sys
from dataclasses import dataclass
from pathlib import Path

from .config import PreprocessConfig
from .discovery import discover_videos
from .segmenter import VariantSpan
from .splitter import variant_filename, view_filename, write_clip

logger = logging.getLogger(__name__)

_VALID_MODES = {"multiway", "multiview", "both"}


@dataclass(frozen=True)
class ManualCutRow:
    """Một dòng yêu cầu cắt tay đã được chuẩn hóa."""

    video_id: str
    mode: str
    cut_seconds: tuple[float, ...]
    view_cut_seconds: tuple[float, ...]
    notes: str = ""


@dataclass(frozen=True)
class PlannedClip:
    """Một clip dự kiến ghi ra: tên file + khoảng frame (inclusive)."""

    out_filename: str
    span: VariantSpan


# --------------------------------------------------------------------------- #
# Đọc bảng (THUẦN trừ lúc mở file)
# --------------------------------------------------------------------------- #
def _parse_seconds(raw: "str | None") -> tuple[float, ...]:
    """Phân tách chuỗi ``"2.5,5.0"`` / ``"2.5;5.0"`` thành tuple float tăng dần."""
    if raw is None:
        return ()
    text = str(raw).strip()
    if not text:
        return ()
    parts = [p for chunk in text.split(";") for p in chunk.split(",")]
    values: list[float] = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        values.append(float(p))
    return tuple(sorted(values))


def _normalize_header(name: str) -> str:
    return str(name).strip().lower().replace(" ", "_")


def parse_manual_rows(records: "list[dict[str, str]]") -> list[ManualCutRow]:
    """Chuyển các dòng dict (đã chuẩn hóa header) thành :class:`ManualCutRow`.

    Bỏ qua (kèm cảnh báo) các dòng thiếu ``video_id`` hoặc ``mode`` không hợp lệ.
    """
    rows: list[ManualCutRow] = []
    for index, rec in enumerate(records):
        norm = {_normalize_header(k): v for k, v in rec.items()}
        video_id = str(norm.get("video_id", "")).strip()
        mode = str(norm.get("mode", "")).strip().lower()
        if not video_id:
            logger.warning("manual_cut: dòng %d thiếu video_id — bỏ qua.", index)
            continue
        if mode not in _VALID_MODES:
            logger.warning(
                "manual_cut: dòng %d (%s) mode không hợp lệ %r — bỏ qua.",
                index,
                video_id,
                mode,
            )
            continue
        try:
            cut = _parse_seconds(norm.get("cut_seconds"))
            view_cut = _parse_seconds(norm.get("view_cut_seconds"))
        except ValueError as exc:
            logger.warning(
                "manual_cut: dòng %d (%s) mốc thời gian không đọc được (%s) — bỏ qua.",
                index,
                video_id,
                exc,
            )
            continue
        rows.append(
            ManualCutRow(
                video_id=video_id,
                mode=mode,
                cut_seconds=cut,
                view_cut_seconds=view_cut,
                notes=str(norm.get("notes", "")).strip(),
            )
        )
    return rows


def read_manual_table(path: Path) -> list[ManualCutRow]:
    """Đọc bảng cắt tay từ ``.csv`` hoặc ``.xlsx`` thành danh sách :class:`ManualCutRow`."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with open(path, "r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.DictReader(fh)
            records = [dict(r) for r in reader]
    elif suffix in (".xlsx", ".xlsm"):
        records = _read_xlsx(path)
    else:
        raise ValueError(f"Định dạng bảng không hỗ trợ: {path.suffix!r} (cần .csv/.xlsx)")
    return parse_manual_rows(records)


def _read_xlsx(path: Path) -> "list[dict[str, str]]":
    """Đọc sheet đầu của XLSX thành list dict {header: value} (dòng đầu là header)."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []
    headers = [str(h) if h is not None else "" for h in header]
    records: list[dict[str, str]] = []
    for values in rows_iter:
        if values is None:
            continue
        rec: dict[str, str] = {}
        for h, v in zip(headers, values):
            rec[h] = "" if v is None else str(v)
        records.append(rec)
    return records


# --------------------------------------------------------------------------- #
# Lập kế hoạch cắt (THUẦN)
# --------------------------------------------------------------------------- #
def _seconds_to_frame(t: float, fps: float, num_frames: int) -> int:
    """Đổi mốc giây → chỉ số frame (0-based), chặn trong [0, num_frames-1]."""
    f = int(round(float(t) * float(fps)))
    return max(0, min(f, num_frames - 1))


def _method_spans(
    cut_seconds: "tuple[float, ...]", fps: float, num_frames: int
) -> list[VariantSpan]:
    """Dựng các span cách liên tục từ mốc bắt đầu cách 2,3… (mức frame, exclusive cuối)."""
    boundaries = sorted(
        {_seconds_to_frame(t, fps, num_frames) for t in cut_seconds if t > 0}
    )
    starts = [0] + boundaries
    spans: list[VariantSpan] = []
    for k, start in enumerate(starts):
        end = (num_frames - 1) if k == len(starts) - 1 else boundaries[k] - 1
        spans.append(VariantSpan(variant_index=k + 1, start_frame=start, end_frame=end))
    return spans


def plan_manual_clips(
    row: ManualCutRow, fps: float, num_frames: int
) -> list[PlannedClip]:
    """Lập danh sách clip (tên + span) cho một dòng cắt tay (THUẦN — không I/O).

    Trả về danh sách rỗng (kèm cảnh báo) nếu yêu cầu không hợp lệ (vd mốc nằm
    ngoài video, span suy biến).
    """
    vid = row.video_id

    if row.mode == "multiway":
        spans = _method_spans(row.cut_seconds, fps, num_frames)
        total = len(spans)
        if total < 2:
            logger.warning(
                "manual_cut[%s]: multiway cần >= 1 mốc cut_seconds — bỏ qua.", vid
            )
            return []
        return [
            PlannedClip(variant_filename(vid, s.variant_index, total), s) for s in spans
        ]

    if row.mode == "multiview":
        if len(row.view_cut_seconds) != 1:
            logger.warning(
                "manual_cut[%s]: multiview cần ĐÚNG 1 mốc view_cut_seconds — bỏ qua.",
                vid,
            )
            return []
        cut = _seconds_to_frame(row.view_cut_seconds[0], fps, num_frames)
        front = VariantSpan(1, 0, cut - 1)
        side = VariantSpan(2, cut, num_frames - 1)
        if front.end_frame < front.start_frame or side.end_frame < side.start_frame:
            logger.warning("manual_cut[%s]: mốc view suy biến — bỏ qua.", vid)
            return []
        return [
            PlannedClip(view_filename(vid, is_side=False), front),
            PlannedClip(view_filename(vid, is_side=True), side),
        ]

    # mode == "both" (METHOD-MAJOR): cắt cách trước, mỗi cách cắt góc.
    method_spans = _method_spans(row.cut_seconds, fps, num_frames)
    total = len(method_spans)
    if total < 2:
        logger.warning(
            "manual_cut[%s]: both cần >= 1 mốc cut_seconds (cắt cách) — bỏ qua.", vid
        )
        return []
    if len(row.view_cut_seconds) != total:
        logger.warning(
            "manual_cut[%s]: both cần %d mốc view_cut_seconds (một cho mỗi cách) "
            "nhưng có %d — bỏ qua.",
            vid,
            total,
            len(row.view_cut_seconds),
        )
        return []

    clips: list[PlannedClip] = []
    for k, mspan in enumerate(method_spans):
        base = variant_filename(vid, mspan.variant_index, total)[: -len(".mp4")]
        vcut = _seconds_to_frame(row.view_cut_seconds[k], fps, num_frames)
        if not (mspan.start_frame < vcut <= mspan.end_frame):
            logger.warning(
                "manual_cut[%s]: mốc view cách %d (%ss) nằm ngoài đoạn cách "
                "[%d,%d] — bỏ qua dòng.",
                vid,
                k + 1,
                row.view_cut_seconds[k],
                mspan.start_frame,
                mspan.end_frame,
            )
            return []
        front = VariantSpan(mspan.variant_index, mspan.start_frame, vcut - 1)
        side = VariantSpan(mspan.variant_index, vcut, mspan.end_frame)
        clips.append(PlannedClip(view_filename(base, is_side=False), front))
        clips.append(PlannedClip(view_filename(base, is_side=True), side))
    return clips


# --------------------------------------------------------------------------- #
# Điều phối (BIÊN)
# --------------------------------------------------------------------------- #
def run_manual_cut(
    cfg: PreprocessConfig,
    input_path: Path,
    *,
    dry_run: bool = False,
    logger_obj: "logging.Logger | None" = None,
) -> int:
    """Đọc bảng cắt tay và ghi các clip. Trả về số clip đã ghi (hoặc dự kiến)."""
    log = logger_obj if logger_obj is not None else logger

    config_errors = cfg.validate()
    if config_errors:
        for err in config_errors:
            print(f"[config error] {err}", file=sys.stderr)
        return 0

    rows = read_manual_table(input_path)
    if not rows:
        log.warning("manual_cut: bảng %s không có dòng hợp lệ.", input_path)
        return 0

    # Index video nguồn theo video_id để lấy đường dẫn + thuộc tính.
    entries = {e.video_id: e for e in discover_videos(cfg, log)}
    from . import video_probe

    out_dir = cfg.split_output_path
    written = 0
    for row in rows:
        entry = entries.get(row.video_id)
        if entry is None:
            log.warning(
                "manual_cut[%s]: không tìm thấy video nguồn trong %s — bỏ qua.",
                row.video_id,
                ", ".join(cfg.video_search_dirs),
            )
            continue
        props = video_probe.probe(entry.path)
        if props is None:
            log.warning("manual_cut[%s]: không probe được thuộc tính — bỏ qua.", row.video_id)
            continue

        clips = plan_manual_clips(row, props.fps, int(props.num_frames))
        if not clips:
            continue

        for clip in clips:
            if dry_run:
                log.info(
                    "manual_cut[%s] (dry-run): %s frames [%d,%d]",
                    row.video_id,
                    clip.out_filename,
                    clip.span.start_frame,
                    clip.span.end_frame,
                )
                written += 1
                continue
            out_path = out_dir / clip.out_filename
            if write_clip(entry.path, out_path, clip.span, props, log):
                written += 1
            else:
                log.error("manual_cut[%s]: ghi clip %s thất bại.", row.video_id, clip.out_filename)

    log.info("manual_cut: %s %d clip.", "dự kiến" if dry_run else "đã ghi", written)
    return written


def _build_config_from_args(args) -> PreprocessConfig:
    project_root = Path(args.project_root) if args.project_root else Path.cwd()
    return PreprocessConfig(project_root=project_root)


def main(argv=None) -> int:
    """Điểm vào CLI. Trả về mã thoát (0 = thành công)."""
    parser = argparse.ArgumentParser(
        prog="qipedc_video_preprocess.manual_cut",
        description="Cắt video theo mốc tự nhập (CSV/XLSX) cho nhiều cách / nhiều góc.",
    )
    parser.add_argument(
        "--input",
        required=True,
        help="Đường dẫn bảng CSV/XLSX chứa các dòng cắt tay.",
    )
    parser.add_argument(
        "--project-root",
        default=None,
        help="Thư mục gốc dự án (mặc định: thư mục hiện tại).",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Chỉ in kế hoạch cắt, không ghi clip.",
    )
    args = parser.parse_args(argv)

    cfg = _build_config_from_args(args)
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    count = run_manual_cut(cfg, Path(args.input), dry_run=args.dry_run)
    print(f"Done. clips={count}")
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
