"""Unit tests cho ``qipedc_video_preprocess.label_writer`` — lớp I/O biên (Req 7).

Phủ các nhánh I/O không thuộc lớp logic thuần (đã được property-test 11/12/13 ở
``test_preprocess_properties.py``):

* **Req 7.4** — :func:`write_new_labels` ghi ``.xlsx`` mới tại
  ``cfg.new_labels_path`` và **không** ghi đè/sửa Bảng_Nhãn_Nguồn trong
  ``Dataset/labels/``.
* **Req 7.7** — khi không có dòng nhãn nào, vẫn ghi file chỉ gồm **dòng header**
  với đúng cấu trúc cột và ghi log.

Dùng ``tmp_path`` của pytest làm ``project_root`` (test read/write, không liên
quan ràng buộc ổ ``D:`` đã được Property 1 phủ riêng).
"""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl

from qipedc_video_preprocess.config import PreprocessConfig
from qipedc_video_preprocess.label_writer import (
    COLUMN_HEADERS,
    LabelRow,
    read_source_labels,
    write_new_labels,
)


class CapturingHandler(logging.Handler):
    def __init__(self) -> None:
        super().__init__(level=logging.DEBUG)
        self.records: list[logging.LogRecord] = []

    def emit(self, record: logging.LogRecord) -> None:
        self.records.append(record)

    def messages(self) -> list[str]:
        return [r.getMessage() for r in self.records]


def _logger() -> tuple[logging.Logger, CapturingHandler]:
    handler = CapturingHandler()
    logger = logging.getLogger(f"test_label_writer_unit.{id(handler)}")
    logger.handlers.clear()
    logger.addHandler(handler)
    logger.setLevel(logging.DEBUG)
    logger.propagate = False
    return logger, handler


def _make_source_workbook(path: Path, rows: list[tuple]) -> None:
    """Ghi một bảng nhãn nguồn (header + rows) ra ``path`` bằng openpyxl."""
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(list(COLUMN_HEADERS))
    for row in rows:
        worksheet.append(list(row))
    workbook.save(str(path))
    workbook.close()


def _cfg(tmp_path: Path) -> PreprocessConfig:
    return PreprocessConfig(
        project_root=tmp_path,
        labels_glob="Dataset/labels/*.xlsx",
        new_labels_path="Dataset/processed_videos/split_variants/labels_split.xlsx",
    )


# --------------------------------------------------------------------------- #
# Req 7.4 — không ghi đè bảng nhãn nguồn
# --------------------------------------------------------------------------- #
def test_write_new_labels_does_not_overwrite_source(tmp_path):
    cfg = _cfg(tmp_path)
    source = tmp_path / "Dataset" / "labels" / "batch_1.xlsx"
    _make_source_workbook(
        source,
        rows=[("1", "D0530", "W00202.mp4", "bắt tay", "Bắc", "chào hỏi", "s1")],
    )
    source_bytes_before = source.read_bytes()

    new_rows = [
        LabelRow(
            stt="1",
            id="D0530",
            video="W00202_c1.mp4",
            label="bắt tay",
            region="Bắc",
            topic="chào hỏi",
            signer="s1",
        )
    ]
    logger, _ = _logger()
    out_path = write_new_labels(new_rows, cfg, logger=logger)

    # Output is a different file under split_variants/ (Req 7.4).
    assert out_path.exists()
    assert out_path != source
    assert out_path.name == "labels_split.xlsx"

    # The source spreadsheet is byte-for-byte unchanged.
    assert source.read_bytes() == source_bytes_before

    # The new file genuinely contains the new VIDEO name.
    wb = openpyxl.load_workbook(str(out_path), read_only=True)
    values = [tuple(r) for r in wb.active.iter_rows(values_only=True)]
    wb.close()
    assert values[0] == COLUMN_HEADERS
    assert any("W00202_c1.mp4" in str(cell) for row in values[1:] for cell in row)


# --------------------------------------------------------------------------- #
# Req 7.7 — không có dòng dữ liệu → chỉ ghi header + log
# --------------------------------------------------------------------------- #
def test_write_new_labels_empty_writes_header_only_and_logs(tmp_path):
    cfg = _cfg(tmp_path)
    logger, handler = _logger()

    out_path = write_new_labels([], cfg, logger=logger)

    assert out_path.exists()
    wb = openpyxl.load_workbook(str(out_path), read_only=True)
    values = [tuple(r) for r in wb.active.iter_rows(values_only=True)]
    wb.close()

    # Exactly one row — the header — with the canonical column structure.
    assert len(values) == 1
    assert values[0] == COLUMN_HEADERS

    # A log entry records that no label rows were produced (Req 7.7).
    assert any("header" in m.lower() or "không có dòng" in m for m in handler.messages())


# --------------------------------------------------------------------------- #
# read_source_labels — header khớp không phân biệt hoa/thường
# --------------------------------------------------------------------------- #
def test_read_source_labels_case_insensitive_header(tmp_path):
    cfg = _cfg(tmp_path)
    source = tmp_path / "Dataset" / "labels" / "weird_case.xlsx"
    source.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    # Header with mixed/odd casing — must still be matched.
    ws.append(["stt", "Id", "ViDeO", "label", "Region", "TOPIC", "Signer"])
    ws.append(["1", "D0530", "W00202.mp4", "bắt tay", "Bắc", "chào hỏi", "s1"])
    wb.save(str(source))
    wb.close()

    rows = read_source_labels(cfg)
    assert len(rows) == 1
    row = rows[0]
    assert row.video == "W00202.mp4"
    assert row.label == "bắt tay"
    assert row.signer == "s1"
